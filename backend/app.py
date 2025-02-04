from flask import Flask, request, send_file
import pandas as pd
import random
from openpyxl import Workbook
from openpyxl.styles import Font
import os
from io import BytesIO

app = Flask(__name__)

def shuffle_students(file):
    xls = pd.ExcelFile(file, engine='openpyxl')
    
    # Define group names
    group_names = ["EEE", "MECH EB", "ECA CSB", "CSA ECB"]
    
    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Process each class in each year
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
        
        # Process each class separately
        for col in df.columns:
            class_name = f"{sheet}_{col}"
            students = df[col].dropna().tolist()
            
            # Shuffle students randomly
            random.shuffle(students)
            
            # Split into 4 groups + overflow
            num_groups = len(group_names)
            group_size = len(students) // num_groups
            groups = {group_names[i]: students[i * group_size: (i + 1) * group_size] for i in range(num_groups)}
            groups['Overflow'] = students[num_groups * group_size:]
            
            # Create a new worksheet for the class
            ws = wb.create_sheet(title=class_name)
            
            # Write groups to sheet
            max_len = max(len(students) for students in groups.values())
            headers = list(groups.keys())
            ws.append(headers)
            
            # Make group names bold
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num).font = Font(bold=True)
            
            for i in range(max_len):
                row = [groups[group][i] if i < len(groups[group]) else "" for group in headers]
                ws.append(row)
    
    # Save workbook to a BytesIO stream and return it
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Rewind the file pointer to the start
    return output

@app.route('/assign-groups', methods=['POST'])
def shuffle():
    if 'file' not in request.files:
        return "No file part", 400
    file = request.files['file']
    if file.filename == '':
        return "No selected file", 400

    # Shuffle students and get the shuffled file in memory
    output = shuffle_students(file)
    
    # Send the shuffled file as a downloadable response
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name="shuffled_students.xlsx")

if __name__ == '__main__':
    app.run(debug=True)